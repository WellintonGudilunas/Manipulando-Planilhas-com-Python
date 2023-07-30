import os
from openpyxl import load_workbook, Workbook

def processar_planilha(planilha):
    dados = {}
    maior_quantidade = 0
    menor_quantidade = None
    maior_receita = 0
    item_maior_quantidade = None
    item_menor_quantidade = None
    item_maior_receita = None

    for linha in planilha.iter_rows(min_row=2, values_only=True):
        print(linha)
        codigo, nome, quantidade, preco = linha

        if quantidade is None:
            quantidade = 0

        if preco is None:
            preco = 0

        # Soma da quantidade de cada item
        if codigo not in dados:
            dados[codigo] = {
                'nome': nome,
                'quantidade_total': 0,
                'receita_total': 0
            }
        dados[codigo]['quantidade_total'] += quantidade
        dados[codigo]['receita_total'] += preco * quantidade

        # Atualizar os itens com maior e menor quantidade
        if quantidade > maior_quantidade:
            maior_quantidade = quantidade
            item_maior_quantidade = codigo

        if menor_quantidade is None or quantidade < menor_quantidade:
            menor_quantidade = quantidade
            item_menor_quantidade = codigo

        # Atualizar o item que gerou maior receita
        receita_item = preco * quantidade
        if receita_item > maior_receita:
            maior_receita = receita_item
            item_maior_receita = codigo

    return dados, item_maior_quantidade, item_menor_quantidade, item_maior_receita

def criar_relatorio(dados, item_maior_quantidade, item_menor_quantidade, item_maior_receita):
    relatorio = Workbook()
    planilha_relatorio = relatorio.active
    planilha_relatorio.append(['Código do Item', 'Nome do Item', 'Quantidade Total', 'Receita Total'])

    quantidade_total_relatorio = 0
    for codigo_item, info in dados.items():
        planilha_relatorio.append([codigo_item, info['nome'], info['quantidade_total'], info['receita_total']])
        quantidade_total_relatorio += info['quantidade_total']  # Somar a quantidade total

    planilha_relatorio.append([])

    # Informações adicionais no relatório
    planilha_relatorio.append(['Item com Maior Quantidade', 'Item com Menor Quantidade', 'Item que Gerou mais Receita'])
    planilha_relatorio.append([item_maior_quantidade, item_menor_quantidade, item_maior_receita])

    planilha_relatorio.append([])  # Linha em branco para separar

    # Adicionar a quantidade total no relatório
    planilha_relatorio.append(['Quantidade Total do Relatório', quantidade_total_relatorio])

    return relatorio

if __name__ == "__main__":
    # Defina o caminho da pasta que contém as planilhas
    pasta_planilhas = 'C:\\Users\\Well\\Desktop\\planilhasExcel\\'

    # Variáveis para armazenar informações de maior e menor quantidade e maior receita
    item_maior_quantidade_total = None
    item_menor_quantidade_total = None
    item_maior_receita_total = None
    maior_quantidade_total = 0

    # Lista para armazenar os dados de todas as planilhas
    todos_dados = {}

    # Loop pelas planilhas na pasta
    for filename in os.listdir(pasta_planilhas):
        if filename.endswith('.xlsx'):
            caminho_arquivo = os.path.join(pasta_planilhas, filename)
            planilha = load_workbook(caminho_arquivo).active

            # Processar cada planilha e obter os dados
            dados, item_maior_quantidade, item_menor_quantidade, item_maior_receita = processar_planilha(planilha)

            if item_maior_quantidade_total is None or dados[item_maior_quantidade_total]['quantidade_total'] < dados[item_maior_quantidade]['quantidade_total']:
                item_maior_quantidade_total = item_maior_quantidade

            if item_menor_quantidade_total is None or dados[item_menor_quantidade_total]['quantidade_total'] > dados[item_menor_quantidade]['quantidade_total']:
                item_menor_quantidade_total = item_menor_quantidade

            if item_maior_receita_total is None or dados[item_maior_receita_total]['receita_total'] < dados[item_maior_receita]['receita_total']:
                item_maior_receita_total = item_maior_receita


            # Armazenar os dados em todos_dados
            for codigo_item, info in dados.items():
                if codigo_item not in todos_dados:
                    todos_dados[codigo_item] = {
                        'nome': info['nome'],
                        'quantidade_total': 0,
                        'receita_total': 0
                    }
                todos_dados[codigo_item]['quantidade_total'] += info['quantidade_total']
                todos_dados[codigo_item]['receita_total'] += info['receita_total']

    # Criar o relatório final
    relatorio_final = criar_relatorio(todos_dados, item_maior_quantidade_total, item_menor_quantidade_total, item_maior_receita_total)

    relatorio_final.save('C:\\Users\\Well\\Desktop\\planilhasExcel\\relatorio_final.xlsx')
